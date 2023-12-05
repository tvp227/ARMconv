import os
import json
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

def process_json_file(file_path):
    with open(file_path, 'r') as file:
        data = json.load(file)

    properties = data['resources'][0]['properties']
    display_name = properties.get('displayName', '')
    description = properties.get('description', '')
    severity = properties.get('severity', '')
    enabled = properties.get('enabled', '')
    query = properties.get('query', '')
    query_frequency = properties.get('queryFrequency', '')
    tactics = ', '.join(properties.get('tactics', []))

    return {
        'DisplayName': display_name,
        'Description': description,
        'Severity': severity,
        'Enabled': enabled,
        'Query': query,
        'QueryFrequency': query_frequency,
        'Tactics': tactics
    }

def main():
    folder_path = filedialog.askdirectory(title="Select Folder")
    if not folder_path:
        print("Folder not selected. Exiting.")
        return

    output_file_path = 'output.xlsx'
    wb = Workbook()
    ws = wb.active

    headers = ['DisplayName', 'Description', 'Severity', 'Enabled', 'Query', 'QueryFrequency', 'Tactics']
    ws.append(headers)

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.json'):
                file_path = os.path.join(root, file)
                json_data = process_json_file(file_path)
                ws.append([json_data[header] for header in headers])

    # Create a table
    table = Table(displayName="Table1", ref=f"A1:G{len(ws['A'])}")
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(output_file_path)
    print(f"Data has been written to {output_file_path}")

if __name__ == "__main__":
    main()
