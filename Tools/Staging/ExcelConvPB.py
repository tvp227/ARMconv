import os
import json
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

def process_json_file(file_path):
    with open(file_path, 'r') as file:
        data = json.load(file)

    metadata = data.get('metadata', {})
    title = metadata.get('title', '')
    description = metadata.get('description', '')

    # Extracting connection IDs
    connection_ids = get_connection_ids(data)

    return {
        'Title': title,
        'Description': description,
        'ConnectionIDs': ', '.join(connection_ids)
    }

def get_last_segment(id_path):
    return id_path.split("/")[-1]

def get_connection_ids(json_data):
    connection_ids = []

    def traverse(obj):
        if isinstance(obj, dict):
            if "$connections" in obj:
                value = obj["$connections"].get("value")
                if value and isinstance(value, dict):
                    for connection_key, connection_value in value.items():
                        if isinstance(connection_value, dict) and "id" in connection_value:
                            connection_id = get_last_segment(connection_value["id"])
                            connection_ids.append(connection_id)
            for key, value in obj.items():
                traverse(value)
        elif isinstance(obj, list):
            for item in obj:
                traverse(item)

    traverse(json_data)
    return connection_ids

def main():
    folder_path = filedialog.askdirectory(title="Select Folder")
    if not folder_path:
        print("Folder not selected. Exiting.")
        return

    output_file_path = 'output.xlsx'
    wb = Workbook()
    ws = wb.active

    headers = ['Title', 'Description', 'ConnectionIDs']
    ws.append(headers)

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.json'):
                file_path = os.path.join(root, file)
                json_data = process_json_file(file_path)
                ws.append([json_data[header] for header in headers])

    # Create a table
    table = Table(displayName="Table1", ref=f"A1:C{len(ws['A'])}")
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(output_file_path)
    print(f"Data has been written to {output_file_path}")

if __name__ == "__main__":
    main()
