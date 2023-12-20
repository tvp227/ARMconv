import os
import json
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

def process_json_file(file_path):
    with open(file_path, 'r') as file:
        data = json.load(file)

    if isinstance(data, list):
        # If the top-level structure is a list, assume the first item is the relevant data
        data = data[0]

    parameters = data.get('parameters', {})
    workbook_name = parameters.get('workbook-name', {}).get('defaultValue', '')

    return {
        'WorkbookName': workbook_name
    }

def main():
    folder_path = filedialog.askdirectory(title="Select Folder")
    if not folder_path:
        print("Folder not selected. Exiting.")
        return

    output_file_path = 'workbooks_output.xlsx'
    wb = Workbook()
    ws = wb.active

    headers = ['WorkbookName']
    ws.append(headers)

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for root, dirs, files in os.walk(folder_path):
        for file in files:
            if file.endswith('.json'):
                file_path = os.path.join(root, file)
                json_data = process_json_file(file_path)
                ws.append([json_data[header] for header in headers])

    # Create a table
    table = Table(displayName="Table1", ref=f"A1:A{len(ws['A'])}")
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(output_file_path)
    print(f"Data has been written to {output_file_path}")

if __name__ == "__main__":
    main()
